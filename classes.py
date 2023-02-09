from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Font, Alignment


class LeitorAcoes:
    def __init__(self, caminho_arquivo: str = ""):
        self.caminho_arquivo = caminho_arquivo
        self.dados = []

    def processa_arquivo(self, acao: str):
        with open(f'{self.caminho_arquivo}{acao}.txt', 'r') as arquivo_cotacao:
            linhas = arquivo_cotacao.readlines()
            self.dados = [linha.replace("\n", "").split(";") for linha in linhas]


class GerenciadorPlanilha:
    def __init__(self):
        self.workbook = Workbook()
        self.planilha_ativa = None

    def adiciona_planilha(self, titulo_planilha: str = ""):
        nova_planilha = self.workbook.create_sheet(titulo_planilha)
        self.workbook.active = nova_planilha
        self.planilha_ativa = nova_planilha

        return nova_planilha

    def adiciona_linas(self, dados: list):
        self.planilha_ativa.append(dados)

    def atualiza_celula(self, celula: str, dado):
        self.planilha_ativa[celula] = dado

    def mescla_celula(self, celula_inicio: str, celula_fim: str):
        self.planilha_ativa.merge.cells(f'{celula_inicio}:{celula_fim}')

    def estiliza_fonte(self, celula: str, fonte: Font):
        self.planilha_ativa[celula].font = fonte
