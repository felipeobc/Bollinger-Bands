from datetime import date
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment

# acao = input("QUal o codigo da Ação que você quer processar: ").upper()

acao = "BIDI4"

with open(f'./dados/{acao}.txt', 'r') as arquivo_cotacao:
    linhas = arquivo_cotacao.readlines()
    linhas = [linha.replace("\n", "").split(";") for linha in linhas]

workbook = Workbook()
planilha_ativa = workbook.active
planilha_ativa.title = "Dados"

planilha_ativa.append(["DATA", "COTAÇÂO", "BANDA INFERIOR", "BANDA SUPERIOR"])
indice = 2

for linha in linhas:
    # DATA
    # 2018-05-10 21:00:00;0.9969
    ano_mes_dia = linha[0].split(" ")[0]
    data = date(
        year=int(ano_mes_dia.split("-")[0]),
        month=int(ano_mes_dia.split("-")[1]),
        day=int(ano_mes_dia.split("-")[2])
    )
    # COTAÇÃO
    cotacao = float(linha[1])

    # Atualiza celulas
    planilha_ativa[f'A{indice}'] = data
    planilha_ativa[f'B{indice}'] = cotacao
    # BANDA INFERIOR
    planilha_ativa[f'C{indice}'] = f'AVERANGE(B{indice}:B{indice + 19}) - 2*STDEV(B{indice}:B{indice + 19})'
    # BANDA SUPERIOR
    planilha_ativa[f'D{indice}'] = f'AVERANGE(B{indice}:B{indice + 19}) + 2*STDEV(B{indice}:B{indice + 19})'

    indice += 1

planilha_grafico = workbook.create_sheet("Grafico")

workbook.active = planilha_grafico

# Configuração de  planilha
planilha_grafico.merge_cells("A1:T2")
cabecalho = planilha_grafico['A1']
cabecalho.font = Font(b=True, sz=18, color="FFFFFF")
cabecalho.fill = PatternFill("solid", fgColor="07838F")
cabecalho.alignment = Alignment(vertical="center", horizontal="center")
cabecalho.value = "Historico de Cotação"

grafico = LineChart()
grafico.title = f'Cotações - {acao}'
grafico.x_axis.title = "Data da Cotação"
grafico.y_axis.title = "Valor da Cotação"

referencia_cotacao = Reference(planilha_ativa, min_col=2, min_row=2, max_col=4, max_row=indice)
referencia_datas = Reference(planilha_ativa, min_col=1, min_row=2, max_col=1, max_row=indice)

grafico.add_data(referencia_cotacao)
grafico.set_categories(referencia_datas)

linha_cotacao = grafico.series[0]
linha_bb_inferior = grafico.series[1]
linha_bb_superior = grafico.series[2]

linha_cotacao.graphicalProperties.line.width = 0
linha_cotacao.graphicalProperties.line.solidFill = "0a55ab"

linha_bb_inferior.graphicalProperties.line.width = 0
linha_bb_inferior.graphicalProperties.line.solidFill = "a61588"

linha_bb_superior.graphicalProperties.line.width = 0
linha_bb_superior.graphicalProperties.line.solidFill = "12a154"

planilha_grafico.add_chart(grafico, "A3")

imagem = Image('./recursos/b3.png')
planilha_grafico.merge_cells("I32:L35")
planilha_grafico.add_image(imagem, "I32")

workbook.save("./saida/Planilha.xlsx")
